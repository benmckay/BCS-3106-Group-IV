from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.exceptions import ValidationError, NotFound, NotAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.views.decorators.http import require_GET
from rest_framework.filters import SearchFilter, OrderingFilter
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Sum, Count, Q, Avg, F, ExpressionWrapper, DecimalField, DurationField
from django.db.models.functions import Coalesce
from django.shortcuts import render
from django.http import HttpResponse
from datetime import timedelta, date
from decimal import Decimal
from collections import defaultdict
import io
import base64
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import numpy as np
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    A4 = landscape = canvas = ImageReader = colors = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    Workbook = Font = Alignment = get_column_letter = XLImage = None

from .models import (
    Customer, Worker, Estimate, Job, Supplier,
    Material, Invoice, Payment
)
from .serializers import (
    UserSerializer, UserRegistrationSerializer,
    CustomerSerializer, CustomerDetailSerializer,
    WorkerSerializer, EstimateSerializer, JobSerializer,
    JobDetailSerializer, SupplierSerializer, MaterialSerializer,
    InvoiceSerializer, InvoiceDetailSerializer, PaymentSerializer
)

PRIMARY_COLOR = '#1b5e20'
SECONDARY_COLOR = '#2e7d32'
ACCENT_COLOR = '#a5d6a7'
CHART_COLORS = ['#e53935', '#43a047', '#8e24aa', '#1e88e5', '#fb8c00']
FAVICON_BYTES = base64.b64decode('iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAOim7xkAAAAASUVORK5CYII=')


def favicon(request):
    response = HttpResponse(FAVICON_BYTES, content_type='image/png')
    response['Cache-Control'] = 'public, max-age=86400'
    return response


def _palette(count, values=None, cmap_name=None):
    if count <= 0:
        return []
    if not CHART_COLORS:
        return []
    colors = []
    for idx in range(count):
        colors.append(CHART_COLORS[idx % len(CHART_COLORS)])
    return colors


def _encode_figure(fig):
    buffer = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buffer, format='png', bbox_inches='tight')
    buffer.seek(0)
    image = base64.b64encode(buffer.read()).decode('utf-8')
    buffer.close()
    plt.close(fig)
    return image


def _month_sequence(count):
    anchor = timezone.now().date().replace(day=1)
    months = []
    year = anchor.year
    month = anchor.month
    for _ in range(count):
        months.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    months.reverse()
    return months


def get_dashboard_stats():
    today = timezone.now().date()
    jobs = Job.objects.all()
    invoices = Invoice.objects.select_related('customer')
    workers = Worker.objects.select_related('user')
    materials = Material.objects.all()
    total_jobs = jobs.count()
    active_jobs = jobs.filter(status='IN_PROGRESS').count()
    scheduled_jobs = jobs.filter(status__in=['SCHEDULED', 'CONFIRMED']).count()
    completed_jobs = jobs.filter(status='COMPLETED').count()
    pending_estimates = Estimate.objects.filter(status='PENDING').count()
    accepted_estimates = Estimate.objects.filter(status='ACCEPTED').count()
    paid_invoices_qs = invoices.filter(status='PAID')
    overdue_invoices = invoices.filter(Q(status='SENT') | Q(status='OVERDUE'), due_date__lt=today).count()
    total_revenue = sum((invoice.total_amount for invoice in paid_invoices_qs), Decimal('0'))
    pending_revenue = Decimal('0')
    for invoice in invoices.exclude(status='PAID'):
        pending_revenue += invoice.total_amount - invoice.amount_paid
    worker_total = workers.count()
    worker_available = workers.filter(is_available=True).count()
    worker_availability = round((worker_available / worker_total) * 100, 2) if worker_total else 0
    duration_qs = jobs.annotate(duration=ExpressionWrapper(F('scheduled_end_date') - F('scheduled_start_date'), output_field=DurationField()))
    avg_duration = duration_qs.aggregate(value=Avg('duration'))['value']
    average_job_duration = avg_duration.days if avg_duration else 0
    customer_satisfaction = round((completed_jobs / total_jobs) * 100, 2) if total_jobs else 0
    material_total = materials.annotate(
        total_cost=ExpressionWrapper(F('quantity') * F('unit_cost'), output_field=DecimalField(max_digits=14, decimal_places=2))
    ).aggregate(total=Coalesce(Sum('total_cost'), Decimal('0')))['total']
    recent_activity = []
    for job in jobs.order_by('-updated_at')[:5]:
        recent_activity.append({
            'type': 'Job',
            'title': job.job_title,
            'status': job.status,
            'timestamp': job.updated_at.isoformat()
        })
    for invoice in invoices.order_by('-updated_at')[:5]:
        recent_activity.append({
            'type': 'Invoice',
            'title': invoice.invoice_number,
            'status': invoice.status,
            'timestamp': invoice.updated_at.isoformat()
        })
    recent_activity = sorted(recent_activity, key=lambda item: item['timestamp'], reverse=True)[:6]
    return {
        'active_jobs': active_jobs,
        'scheduled_jobs': scheduled_jobs,
        'completed_jobs': completed_jobs,
        'pending_estimates': pending_estimates,
        'accepted_estimates': accepted_estimates,
        'paid_invoices': paid_invoices_qs.count(),
        'overdue_invoices': overdue_invoices,
        'total_revenue': float(total_revenue),
        'pending_revenue': float(pending_revenue),
        'worker_availability': worker_availability,
        'worker_counts': {
            'total': worker_total,
            'available': worker_available
        },
        'material_spend': float(material_total or 0),
        'average_job_duration': average_job_duration,
        'customer_satisfaction': customer_satisfaction,
        'recent_activity': recent_activity,
        'last_updated': timezone.now().isoformat()
    }


def chart_job_status():
    data = list(Job.objects.values('status').annotate(total=Count('id')).order_by('status'))
    if not data:
        return None
    labels = [dict(Job.STATUS_CHOICES).get(item['status'], item['status']) for item in data]
    totals = [item['total'] for item in data]
    fig, ax = plt.subplots(figsize=(6, 6))
    colors = _palette(len(labels), totals, 'Greens')
    ax.pie(totals, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors)
    ax.set_title('Job Status Distribution', color=PRIMARY_COLOR)
    return {'key': 'job_status', 'title': 'Job Status Distribution', 'image': _encode_figure(fig)}


def chart_invoice_status():
    data = list(Invoice.objects.values('status').annotate(total=Count('id')).order_by('status'))
    if not data:
        return None
    labels = [dict(Invoice.STATUS_CHOICES).get(item['status'], item['status']) for item in data]
    totals = [item['total'] for item in data]
    fig, ax = plt.subplots(figsize=(8, 5))
    colors = _palette(len(labels), totals, 'YlOrBr')
    ax.bar(labels, totals, color=colors)
    ax.set_title('Invoice Status Distribution', color=PRIMARY_COLOR)
    ax.set_ylabel('Invoices')
    plt.xticks(rotation=30, ha='right')
    return {'key': 'invoice_status', 'title': 'Invoice Status Distribution', 'image': _encode_figure(fig)}


def chart_revenue_trend():
    start_date = timezone.now().date() - timedelta(days=180)
    invoices = Invoice.objects.filter(invoice_date__gte=start_date).order_by('invoice_date')
    if not invoices:
        return None
    monthly_totals = defaultdict(Decimal)
    for invoice in invoices:
        month_key = invoice.invoice_date.strftime('%b %Y')
        monthly_totals[month_key] += invoice.total_amount
    months = list(monthly_totals.keys())
    values = [float(monthly_totals[month]) for month in months]
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(months, values, color=CHART_COLORS[3], linewidth=3, marker='o')
    ax.set_title('Revenue Trend (6 Months)', color=PRIMARY_COLOR)
    ax.set_ylabel('Revenue ($)')
    ax.set_xlabel('Month')
    ax.grid(alpha=0.3)
    plt.xticks(rotation=35, ha='right')
    return {'key': 'revenue_trend', 'title': 'Revenue Trend (6 Months)', 'image': _encode_figure(fig)}


def chart_top_materials():
    materials = list(
        Material.objects.annotate(
            total_cost=ExpressionWrapper(F('quantity') * F('unit_cost'), output_field=DecimalField(max_digits=14, decimal_places=2))
        ).order_by('-total_cost')[:10]
    )
    if not materials:
        return None
    names = [material.name for material in materials][::-1]
    costs = [float(material.total_cost) for material in materials][::-1]
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = _palette(len(names), costs, 'YlGn')
    ax.barh(names, costs, color=colors)
    ax.set_title('Top Materials by Cost', color=PRIMARY_COLOR)
    ax.set_xlabel('Cost ($)')
    return {'key': 'materials_cost', 'title': 'Top Materials by Cost', 'image': _encode_figure(fig)}


def chart_worker_cost_breakdown():
    data = list(Worker.objects.values('worker_type').annotate(total=Sum('hourly_rate')).order_by('worker_type'))
    if not data:
        return None
    labels = [dict(Worker.WORKER_TYPES).get(item['worker_type'], item['worker_type']) for item in data]
    totals = [float(item['total']) for item in data]
    fig, ax = plt.subplots(figsize=(8, 5))
    colors = _palette(len(labels), totals, 'BuPu')
    ax.bar(labels, totals, color=colors)
    ax.set_title('Cost Breakdown by Worker Type', color=PRIMARY_COLOR)
    ax.set_ylabel('Hourly Rate Total ($)')
    plt.xticks(rotation=35, ha='right')
    return {'key': 'worker_costs', 'title': 'Cost Breakdown by Worker Type', 'image': _encode_figure(fig)}


def chart_worker_distribution():
    data = list(Worker.objects.values('worker_type').annotate(total=Count('id')).order_by('worker_type'))
    if not data:
        return None
    labels = [dict(Worker.WORKER_TYPES).get(item['worker_type'], item['worker_type']) for item in data]
    totals = [item['total'] for item in data]
    fig, ax = plt.subplots(figsize=(6, 6))
    colors = _palette(len(labels), totals, 'PuBuGn')
    ax.pie(totals, labels=labels, autopct='%1.1f%%', colors=colors)
    ax.set_title('Worker Type Distribution', color=PRIMARY_COLOR)
    return {'key': 'worker_distribution', 'title': 'Worker Type Distribution', 'image': _encode_figure(fig)}


def chart_worker_productivity():
    entries = []
    workers = Worker.objects.select_related('user')
    for worker in workers:
        completed = worker.jobs.filter(status='COMPLETED').count()
        scheduled = worker.jobs.filter(status__in=['SCHEDULED', 'CONFIRMED']).count()
        if completed == 0 and scheduled == 0:
            continue
        name = worker.user.get_full_name() or worker.user.username
        earnings = float(worker.hourly_rate) * completed * 8
        entries.append((name, completed, scheduled, earnings))
    entries = sorted(entries, key=lambda item: item[1], reverse=True)[:8]
    if not entries:
        return None
    names = [item[0] for item in entries]
    completed_values = [item[1] for item in entries]
    scheduled_values = [item[2] for item in entries]
    earnings = [item[3] for item in entries]
    x = np.arange(len(names))
    fig, ax1 = plt.subplots(figsize=(12, 6))
    ax1.bar(x - 0.15, completed_values, width=0.3, label='Completed', color=CHART_COLORS[0])
    ax1.bar(x + 0.15, scheduled_values, width=0.3, label='Scheduled', color=CHART_COLORS[1])
    ax1.set_xticks(x)
    ax1.set_xticklabels(names, rotation=35, ha='right')
    ax1.set_ylabel('Jobs')
    ax1.legend(loc='upper left')
    ax2 = ax1.twinx()
    ax2.plot(x, earnings, color=CHART_COLORS[2], linewidth=3, marker='o', label='Earnings')
    ax2.set_ylabel('Earnings ($)')
    ax2.legend(loc='upper right')
    ax1.set_title('Worker Productivity Metrics', color=PRIMARY_COLOR)
    return {'key': 'worker_productivity', 'title': 'Worker Productivity Metrics', 'image': _encode_figure(fig)}


def chart_monthly_completion():
    months = _month_sequence(12)
    labels = []
    totals = []
    for year, month in months:
        label_date = date(year, month, 1)
        labels.append(label_date.strftime('%b %Y'))
        totals.append(
            Job.objects.filter(status='COMPLETED', actual_end_date__year=year, actual_end_date__month=month).count()
        )
    if not any(totals):
        return None
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, totals, color=CHART_COLORS[4], linewidth=3, marker='o')
    ax.set_title('Monthly Job Completion Rate', color=PRIMARY_COLOR)
    ax.set_ylabel('Completed Jobs')
    ax.set_xlabel('Month')
    ax.grid(alpha=0.3)
    plt.xticks(rotation=35, ha='right')
    return {'key': 'monthly_completion', 'title': 'Monthly Job Completion Rate', 'image': _encode_figure(fig)}


def chart_customer_completion():
    customers = list(
        Customer.objects.annotate(
            total_jobs=Count('jobs'),
            completed_jobs=Count('jobs', filter=Q(jobs__status='COMPLETED'))
        ).filter(total_jobs__gt=0).order_by('-completed_jobs')[:8]
    )
    if not customers:
        return None
    names = [customer.full_name for customer in customers]
    rates = [round((customer.completed_jobs / customer.total_jobs) * 100, 2) for customer in customers]
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = _palette(len(names), rates, 'YlGnBu')
    ax.bar(names, rates, color=colors)
    ax.set_title('Customer Completion Rates', color=PRIMARY_COLOR)
    ax.set_ylabel('Completion %')
    plt.xticks(rotation=35, ha='right')
    return {'key': 'customer_completion', 'title': 'Customer Completion Rates', 'image': _encode_figure(fig)}


def generate_dashboard_charts():
    builders = [
        chart_job_status,
        chart_invoice_status,
        chart_revenue_trend,
        chart_top_materials,
        chart_worker_cost_breakdown,
        chart_worker_distribution,
        chart_worker_productivity,
        chart_monthly_completion,
        chart_customer_completion
    ]
    charts = {}
    for builder in builders:
        try:
            chart = builder()
        except Exception:
            chart = None
        if chart:
            charts[chart['key']] = {
                'title': chart['title'],
                'image': chart['image']
            }
    return charts


def build_pdf_report(stats, charts):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="dashboard-report.pdf"'
    pdf = canvas.Canvas(response, pagesize=landscape(A4))
    page_width, page_height = landscape(A4)
    pdf.setFont('Helvetica-Bold', 22)
    pdf.setFillColor(colors.HexColor(PRIMARY_COLOR))
    pdf.drawString(40, page_height - 50, 'Construction Intelligence Dashboard')
    pdf.setFont('Helvetica', 12)
    pdf.setFillColor(colors.black)
    pdf.drawString(40, page_height - 80, f"Generated {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    summary_rows = [
        ('Active Jobs', 'active_jobs'),
        ('Scheduled Jobs', 'scheduled_jobs'),
        ('Completed Jobs', 'completed_jobs'),
        ('Pending Estimates', 'pending_estimates'),
        ('Accepted Estimates', 'accepted_estimates'),
        ('Paid Invoices', 'paid_invoices'),
        ('Overdue Invoices', 'overdue_invoices'),
        ('Total Revenue', 'total_revenue'),
        ('Pending Revenue', 'pending_revenue'),
        ('Material Spend', 'material_spend'),
        ('Average Job Duration (days)', 'average_job_duration'),
        ('Worker Availability (%)', 'worker_availability'),
        ('Customer Satisfaction (%)', 'customer_satisfaction')
    ]
    pdf.setFont('Helvetica', 11)
    y = page_height - 120
    for label, key in summary_rows:
        value = stats.get(key, 0)
        if key in ['total_revenue', 'pending_revenue', 'material_spend']:
            value_text = f"${value:,.2f}"
        elif key in ['worker_availability', 'customer_satisfaction']:
            value_text = f"{value:.1f}%"
        else:
            value_text = f"{value:,}"
        pdf.drawString(40, y, f"{label}: {value_text}")
        y -= 18
        if y < 80:
            pdf.showPage()
            page_width, page_height = landscape(A4)
            pdf.setFont('Helvetica', 11)
            y = page_height - 60
    chart_items = list(charts.values())
    if chart_items:
        pdf.showPage()
        index = 0
        while index < len(chart_items):
            page_width, page_height = landscape(A4)
            x_positions = [40, page_width / 2 + 20]
            y = page_height - 80
            for column in range(2):
                if index >= len(chart_items):
                    break
                chart = chart_items[index]
                pdf.setFont('Helvetica-Bold', 14)
                pdf.setFillColor(colors.black)
                pdf.drawString(x_positions[column], y, chart['title'])
                image_stream = io.BytesIO(base64.b64decode(chart['image']))
                pdf.drawImage(
                    ImageReader(image_stream),
                    x_positions[column],
                    y - 270,
                    width=page_width / 2 - 80,
                    height=220,
                    preserveAspectRatio=True,
                    mask='auto'
                )
                index += 1
            if index < len(chart_items):
                pdf.showPage()
    pdf.save()
    return response


def build_excel_report(stats, charts):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'
    summary_sheet.append(['Metric', 'Value'])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    summary_data = [
        ('Active Jobs', stats.get('active_jobs', 0)),
        ('Scheduled Jobs', stats.get('scheduled_jobs', 0)),
        ('Completed Jobs', stats.get('completed_jobs', 0)),
        ('Pending Estimates', stats.get('pending_estimates', 0)),
        ('Accepted Estimates', stats.get('accepted_estimates', 0)),
        ('Paid Invoices', stats.get('paid_invoices', 0)),
        ('Overdue Invoices', stats.get('overdue_invoices', 0)),
        ('Total Revenue', f"${stats.get('total_revenue', 0):,.2f}"),
        ('Pending Revenue', f"${stats.get('pending_revenue', 0):,.2f}"),
        ('Material Spend', f"${stats.get('material_spend', 0):,.2f}"),
        ('Average Job Duration (days)', stats.get('average_job_duration', 0)),
        ('Worker Availability (%)', f"{stats.get('worker_availability', 0):.1f}%"),
        ('Customer Satisfaction (%)', f"{stats.get('customer_satisfaction', 0):.1f}%")
    ]
    for label, value in summary_data:
        summary_sheet.append([label, value])
    for column in range(1, 3):
        summary_sheet.column_dimensions[get_column_letter(column)].width = 32
    activity_sheet = workbook.create_sheet(title='Recent Activity')
    activity_sheet.append(['Type', 'Title', 'Status', 'Timestamp'])
    for cell in activity_sheet[1]:
        cell.font = Font(bold=True)
    for item in stats.get('recent_activity', []):
        activity_sheet.append([
            item.get('type'),
            item.get('title'),
            item.get('status'),
            item.get('timestamp')
        ])
    for column in range(1, 5):
        activity_sheet.column_dimensions[get_column_letter(column)].width = 30
    existing_titles = {sheet.title for sheet in workbook.worksheets}
    for index, chart in enumerate(charts.values(), start=1):
        base_title = f"Chart {index}"
        title = base_title
        suffix = 1
        while title in existing_titles:
            suffix += 1
            title = f"{base_title} {suffix}"
        chart_sheet = workbook.create_sheet(title=title[:31])
        chart_sheet.append([chart['title']])
        chart_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        chart_sheet['A1'].font = Font(bold=True, size=16)
        chart_sheet['A1'].alignment = Alignment(horizontal='center')
        image_stream = io.BytesIO(base64.b64decode(chart['image']))
        image = XLImage(image_stream)
        image.width = 960
        image.height = 420
        chart_sheet.add_image(image, 'A3')
        for column in range(1, 5):
            chart_sheet.column_dimensions[get_column_letter(column)].width = 35
        existing_titles.add(chart_sheet.title)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="dashboard-report.xlsx"'
    return response


# Authentication Views
@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    """Register a new user"""
    serializer = UserRegistrationSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        return Response({
            'message': 'User registered successfully',
            'user': UserSerializer(user).data
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def current_user(request):
    """Get current logged-in user details"""
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


# ViewSets
class CustomerViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Customer CRUD operations
    Provides list, create, retrieve, update, and delete operations
    """
    queryset = Customer.objects.all()
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name', 'email', 'phone', 'city']
    ordering_fields = ['created_at', 'first_name', 'last_name']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return CustomerDetailSerializer
        return CustomerSerializer
    
    @action(detail=True, methods=['get'])
    def estimates(self, request, pk=None):
        """Get all estimates for a specific customer"""
        customer = self.get_object()
        estimates = customer.estimates.all()
        serializer = EstimateSerializer(estimates, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def jobs(self, request, pk=None):
        """Get all jobs for a specific customer"""
        customer = self.get_object()
        jobs = customer.jobs.all()
        serializer = JobSerializer(jobs, many=True)
        return Response(serializer.data)


class WorkerViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Worker CRUD operations
    """
    queryset = Worker.objects.all()
    serializer_class = WorkerSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['worker_type', 'is_available']
    search_fields = ['user__first_name', 'user__last_name', 'worker_type']
    ordering_fields = ['created_at', 'hourly_rate', 'experience_years']
    ordering = ['user__first_name']
    
    @action(detail=False, methods=['get'])
    def available(self, request):
        """Get all available workers"""
        workers = self.queryset.filter(is_available=True)
        serializer = self.get_serializer(workers, many=True)
        return Response(serializer.data)


class EstimateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Estimate CRUD operations
    """
    queryset = Estimate.objects.all()
    serializer_class = EstimateSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'customer']
    search_fields = ['customer__first_name', 'customer__last_name', 'work_description']
    ordering_fields = ['created_at', 'property_visit_date', 'estimated_cost']
    ordering = ['-created_at']
    
    def perform_create(self, serializer):
        """Set the created_by field to the current user"""
        serializer.save(created_by=self.request.user)
    
    @action(detail=False, methods=['get'])
    def pending_visits(self, request):
        """Get estimates pending property visit"""
        estimates = self.queryset.filter(status='PENDING')
        serializer = self.get_serializer(estimates, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def accepted(self, request):
        """Get accepted estimates ready for job scheduling"""
        estimates = self.queryset.filter(status='ACCEPTED')
        serializer = self.get_serializer(estimates, many=True)
        return Response(serializer.data)


class JobViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Job CRUD operations
    """
    queryset = Job.objects.all()
    serializer_class = JobSerializer
    permission_classes = [AllowAny]  # allow unauthenticated access

    def perform_create(self, serializer):
        # Only set managed_by if user is authenticated
        if self.request.user and self.request.user.is_authenticated:
            serializer.save(managed_by=self.request.user)
        else:
            serializer.save()  # skip managed_by for anonymous users
    
    @action(detail=False, methods=['get'])
    def upcoming(self, request):
        """Get upcoming jobs"""
        upcoming_jobs = self.queryset.filter(
            scheduled_start_date__gte=timezone.now().date(),
            status__in=['SCHEDULED', 'CONFIRMED']
        )
        serializer = self.get_serializer(upcoming_jobs, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def in_progress(self, request):
        """Get jobs currently in progress"""
        in_progress_jobs = self.queryset.filter(status='IN_PROGRESS')
        serializer = self.get_serializer(in_progress_jobs, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def needs_confirmation(self, request):
        """Get jobs that need customer confirmation"""
        today = timezone.now().date()
        confirmation_window = today + timedelta(days=5)
        jobs = self.queryset.filter(
            scheduled_start_date__range=[today + timedelta(days=1), confirmation_window],
            status='SCHEDULED'
        )
        serializer = self.get_serializer(jobs, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def confirm(self, request, pk=None):
        """Confirm a job start date"""
        job = self.get_object()
        job.status = 'CONFIRMED'
        job.confirmation_date = timezone.now().date()
        job.save()
        serializer = self.get_serializer(job)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Start a job"""
        job = self.get_object()
        job.status = 'IN_PROGRESS'
        job.actual_start_date = timezone.now().date()
        job.save()
        serializer = self.get_serializer(job)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Complete a job"""
        job = self.get_object()
        job.status = 'COMPLETED'
        job.actual_end_date = timezone.now().date()
        job.save()
        serializer = self.get_serializer(job)
        return Response(serializer.data)


class SupplierViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Supplier CRUD operations
    """
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'contact_person', 'email']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']


class MaterialViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Material CRUD operations
    """
    queryset = Material.objects.all()
    serializer_class = MaterialSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['job', 'supplier', 'is_delivered']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'order_date', 'expected_delivery_date']
    ordering = ['-created_at']
    
    @action(detail=False, methods=['get'])
    def pending_delivery(self, request):
        """Get materials pending delivery"""
        materials = self.queryset.filter(is_delivered=False, order_date__isnull=False)
        serializer = self.get_serializer(materials, many=True)
        return Response(serializer.data)


class InvoiceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Invoice CRUD operations
    """
    queryset = Invoice.objects.all()
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'customer']
    search_fields = ['invoice_number', 'customer__first_name', 'customer__last_name']
    ordering_fields = ['invoice_date', 'due_date', 'created_at']
    ordering = ['-invoice_date']
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return InvoiceDetailSerializer
        return InvoiceSerializer
    
    @action(detail=False, methods=['get'])
    def overdue(self, request):
        """Get overdue invoices"""
        today = timezone.now().date()
        overdue_invoices = self.queryset.filter(
            Q(status='SENT') | Q(status='OVERDUE'),
            due_date__lt=today
        )
        # Update status to overdue
        overdue_invoices.update(status='OVERDUE')
        serializer = self.get_serializer(overdue_invoices, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def unpaid(self, request):
        """Get all unpaid invoices"""
        unpaid_invoices = self.queryset.filter(status__in=['SENT', 'OVERDUE'])
        serializer = self.get_serializer(unpaid_invoices, many=True)
        return Response(serializer.data)


class PaymentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Payment CRUD operations
    """
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['invoice', 'payment_method']
    search_fields = ['transaction_reference', 'invoice__invoice_number']
    ordering_fields = ['payment_date', 'amount', 'created_at']
    ordering = ['-payment_date']
    
    def perform_create(self, serializer):
        """Set the received_by field to the current user"""
        serializer.save(received_by=self.request.user)


# Dashboard and Analytics Views
@api_view(['GET'])
@permission_classes([AllowAny])
def dashboard_view(request):
    return render(request._request, 'dashboard.html')


@api_view(['GET'])
@permission_classes([AllowAny])
def dashboard_stats(request):
    return Response(get_dashboard_stats())


@api_view(['GET'])
@permission_classes([AllowAny])
def dashboard_charts(request):
    return Response(generate_dashboard_charts())


@api_view(['GET'])
@permission_classes([AllowAny])
def export_dashboard(request):
    export_format = request.query_params.get('format', 'pdf').lower()
    stats = get_dashboard_stats()
    charts = generate_dashboard_charts()
    if export_format == 'excel':
        if not HAS_OPENPYXL:
            return Response({'detail': 'Excel export requires openpyxl. Install it via pip to enable this feature.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return build_excel_report(stats, charts)
    if export_format == 'pdf':
        if not HAS_REPORTLAB:
            return Response({'detail': 'PDF export requires reportlab. Install it via pip to enable this feature.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return build_pdf_report(stats, charts)
    return Response({'detail': 'Unsupported format'}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def reports(request):
    """
    Generate various reports for the system
    """
    report_type = request.query_params.get('type', 'summary')
    
    if report_type == 'summary':
        # Overall system summary
        return Response({
            'total_customers': Customer.objects.count(),
            'total_jobs': Job.objects.count(),
            'total_revenue': float(Invoice.objects.filter(status='PAID').aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0),
            'active_jobs': Job.objects.filter(status='IN_PROGRESS').count(),
            'pending_invoices': Invoice.objects.filter(status__in=['SENT', 'OVERDUE']).count(),
        })
    
    elif report_type == 'customer':
        # Customer report with their jobs and payments
        customers = Customer.objects.all()
        customer_data = []
        for customer in customers:
            customer_data.append({
                'id': customer.id,
                'name': customer.full_name,
                'email': customer.email,
                'total_jobs': customer.jobs.count(),
                'completed_jobs': customer.jobs.filter(status='COMPLETED').count(),
                'total_spent': float(customer.invoices.filter(status='PAID').aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0)
            })
        return Response(customer_data)
    
    elif report_type == 'financial':
        # Financial report
        paid_invoices = Invoice.objects.filter(status='PAID')
        unpaid_invoices = Invoice.objects.filter(status__in=['SENT', 'OVERDUE'])
        
        return Response({
            'total_revenue': float(paid_invoices.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0),
            'pending_revenue': float(unpaid_invoices.count()) * 1000,  # Simplified
            'total_invoices': Invoice.objects.count(),
            'paid_invoices': paid_invoices.count(),
            'unpaid_invoices': unpaid_invoices.count(),
        })
    
    return Response({'error': 'Invalid report type'}, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.db.models import F, Sum, ExpressionWrapper, DecimalField
from .models import Material

class TopMaterialsByCost(APIView):
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        try:
            limit = int(request.query_params.get('limit', 10))
        except (TypeError, ValueError):
            limit = 10
        limit = max(1, min(limit, 50))
        queryset = (
            Material.objects.select_related('supplier', 'job')
            .annotate(
                total_cost_value=ExpressionWrapper(
                    F('quantity') * F('unit_cost'),
                    output_field=DecimalField(max_digits=18, decimal_places=2)
                )
            )
            .order_by('-total_cost_value')[:limit]
        )
        results = []
        for material in queryset:
            total_cost = getattr(material, 'total_cost_value', None)
            if total_cost is None and material.quantity is not None and material.unit_cost is not None:
                total_cost = material.quantity * material.unit_cost
            quantity = float(material.quantity) if material.quantity is not None else 0.0
            unit_cost = float(material.unit_cost) if material.unit_cost is not None else 0.0
            results.append({
                'id': material.id,
                'name': material.name,
                'unit': material.unit,
                'quantity': quantity,
                'unit_cost': unit_cost,
                'total_cost': float(total_cost or (quantity * unit_cost)),
                'supplier': material.supplier.name if material.supplier else None,
                'job_id': material.job_id,
                'job_title': material.job.job_title if hasattr(material, 'job') and material.job else None,
            })
        return Response({
            'count': len(results),
            'results': results,
            'generated_at': timezone.now().isoformat()
        })

    def _normalize_payload(self, payload):
        if isinstance(payload, list):
            if not payload:
                raise ValidationError({'detail': 'Payload list cannot be empty.'})
            return payload
        if isinstance(payload, dict):
            return [payload]
        raise ValidationError({'detail': 'Payload must be an object or list of objects.'})

    def _update_material(self, record, partial, request):
        material_id = record.get('id') if isinstance(record, dict) else None
        if not material_id:
            raise ValidationError({'id': 'Material id is required for updates.'})
        try:
            instance = Material.objects.get(pk=material_id)
        except Material.DoesNotExist:
            raise NotFound(f'Material with id {material_id} not found.')
        serializer = MaterialSerializer(
            instance,
            data=record,
            partial=partial,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return serializer.data

    def put(self, request, *args, **kwargs):
        records = self._normalize_payload(request.data)
        updated = [self._update_material(record, partial=False, request=request) for record in records]
        return Response({'updated': len(updated), 'results': updated})

    def patch(self, request, *args, **kwargs):
        records = self._normalize_payload(request.data)
        updated = [self._update_material(record, partial=True, request=request) for record in records]
        return Response({'updated': len(updated), 'results': updated})
